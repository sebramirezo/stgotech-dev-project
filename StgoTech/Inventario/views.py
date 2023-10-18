from django.shortcuts import redirect, render , get_object_or_404
from django.contrib.auth.decorators import login_required
from django.views.generic import ListView, DetailView
from .forms import *
from django.db.models import Q , Sum
from django.http.response import JsonResponse
import json
from django.core import serializers
from django.contrib.auth import logout
from django.contrib.auth.views import LoginView
from django.contrib.auth.decorators import login_required, permission_required

from django.http import HttpResponse
from django.contrib.auth import get_user_model
from django.contrib.auth.models import User
from openpyxl import Workbook
from openpyxl.styles import Alignment , Font , Border, Side


# Vistas relacionadas al inicio y cierre de sesión
def redirect_login(request):
    return redirect('login')

class CustomLoginView(LoginView):
    template_name = 'registration/login.html'

def cerrar_sesion(request):
    logout(request)
    return redirect('/login')

####################################################


#VISTAS DE INICIO
#REDIRIGE A LA PAGINA PRINCIPAL INICIO
def index(request):
    return redirect('dashboard')






#OBTIENE LOS DATOS QUE RELACIONADOS A LA BUSQUEDA QUE SE REALIZO
# def buscar_datos_inicio(request):
#     draw = int(request.GET.get('draw', 0))
#     start = int(request.GET.get('start', 0))
#     length = int(request.GET.get('length', 10))
#     search_value = request.GET.get('n', '')

#     resultados = []

#     # Consulta en Comat y realiza JOIN con Incoming y Consumos
#     resultados += list(
#         Incoming.objects.filter(
#             Q(part_number__icontains=search_value) | Q(stdf_fk__stdf_pk__icontains=search_value)
#         ).select_related('stdf_fk', 'ubicacion_fk', 'categoria_fk', 'owner_fk').values(
#             "sn_batch_pk",
#             "categoria_fk__name_categoria",
#             "part_number",
#             "categoria_fk__name_categoria",
#             "descripcion",
#             "stdf_fk__stdf_pk",
#             "qty",
#             "saldo",
#             "stdf_fk__awb",
#             "stdf_fk__num_manifiesto",
#             "owner_fk__name_owner",
#             "ubicacion_fk__name_ubicacion",
#             "f_vencimiento"
#         ).annotate(
#             qty_extraida_total=Sum('consumos__qty_extraida')  # Suma de la cantidad extraída
#         ).distinct()  # Utiliza distinct para obtener un registro por número de parte
#         [start:start + length]
#     )

#     records_filtered = len(resultados)

#     return JsonResponse({
#         "data": resultados,
#         "draw": draw,
#         "recordsTotal": records_filtered,  # Total de registros sin filtrar
#         "recordsFiltered": records_filtered  # Total de registros después del filtrado
#     })



#BUSCADOR DE LA PAGINA DE INICIO, REDIRIGE AL RESULTADO DE BUSQUEDA DE LA PAGINA DE INICIO
def buscar_productos_inicio(request):
    # Obtiene el término de búsqueda del usuario desde la URL
    query_inicio = request.GET.get('n', '')

    return render(request, 'resultado_busqueda_inicio.html', {'query_inicio':query_inicio})

def buscar_datos_inicio(request):
    draw = int(request.GET.get('draw', 0))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))
    search_value = request.GET.get('n', '')

    # Consulta en Incoming
    query_incoming = Incoming.objects.filter(
        Q(part_number__icontains=search_value) | Q(stdf_fk__stdf_pk__icontains=search_value)
    ).select_related('stdf_fk', 'ubicacion_fk', 'categoria_fk', 'owner_fk')

    total_records_incoming = query_incoming.count()

    resultados_incoming = list(
        query_incoming.values(
            "sn_batch_pk",
            "categoria_fk__name_categoria",
            "part_number",
            "descripcion",
            "stdf_fk__stdf_pk",
            "qty",
            "saldo",
            "stdf_fk__awb",
            "stdf_fk__num_manifiesto",
            "owner_fk__name_owner",
            "ubicacion_fk__name_ubicacion",
            "f_vencimiento"
        ).annotate(
            qty_extraida_total=Sum('consumos__qty_extraida')
        ).distinct()
    )[start:start + length]  # Ajustar la paginación aquí

    # Consulta para obtener Comats sin Incoming relacionado
    query_comats_sin_incoming = Comat.objects.filter(incoming__isnull=True)
    total_records_comats_sin_incoming = query_comats_sin_incoming.count()
    resultados_comats_sin_incoming = list(
        query_comats_sin_incoming.values("stdf_pk", "awb", "hawb", "prioridad")
    )[start:start + length]  # Ajustar la paginación aquí

    # Crear el diccionario principal "data"
    data = {
        "draw": draw,
        "recordsTotal": total_records_incoming,
        "recordsFiltered": total_records_incoming,
        "data_incoming": resultados_incoming,
        "data_comats_sin_incoming": resultados_comats_sin_incoming,   
    }

    return JsonResponse(data)






#OBTIENE LOS DATOS PARA REALIZAR EL DETALLE DE LA PAGINA DE INICIO
def detalle_inicio(request, sn_batch_pk):

    detalle_inicio = Incoming.objects.get(sn_batch_pk=sn_batch_pk)   

    comat_data = detalle_inicio.stdf_fk

    consumos_data = detalle_inicio.consumos_set.all()

    return render(request,'detalle_inicio.html' , {'detalle_inicio':detalle_inicio, 'comat_data' : comat_data, 'consumos_data':consumos_data })

@login_required
def comat(request):
    get_form_comat = Comat.objects.all()
    total_cif = 0

    if request.method == 'POST':
        form_comat = ComatForm(request.POST)
        if form_comat.is_valid():
            if request.user.is_authenticated:
                # Obtén el objeto Comat sin guardarlo aún
                comat = form_comat.save(commit=False)
                comat.usuario = request.user  

                # Realiza la suma de fob, flete y seguro
                total_cif = comat.fob + comat.flete + comat.seguro

                # Guarda el objeto Comat con la cif actualizada
                comat.sum_cif = total_cif
                comat.save()

                return redirect('/comat')
            else:
                # Manejo del caso en el que el usuario no está autenticado
                return HttpResponse("Debes iniciar sesión para realizar esta acción.")
    else:
        form_comat = ComatForm()

    context = {
        'form_comat': form_comat,
        'get_form_comat': get_form_comat,
    }
    return render(request, 'comat.html', context)

#BUSCADOR DE COMAT
def buscar_productos(request):
    # Obtiene el término de búsqueda del usuario desde la URL
    query_comat = request.GET.get('c', '')

    return render(request, 'resultado_busqueda_stdf.html', {'query_comat':query_comat})
    #resultados = Comat.objects.filter(id_stdf__icontains=query)
    #resultados_awb = Comat.objects.filter(awb__icontains=query_awb)

    
    #'resultados_awb': resultados_awb, 'query': query, 'query_awb': query_awb



# def obtener_datos_comat(request):
#     # Obtén los parámetros enviados por DataTables
#     draw = int(request.GET.get('draw', 0))
#     start = int(request.GET.get('start', 0))
#     length = int(request.GET.get('length', 10))  # Número de registros por página

#     # Realiza la consulta teniendo en cuenta la paginación
#     comat_data = Comat.objects.all()[start:start + length]

#     # Formatea los datos en un formato compatible con DataTables
#     data = []
#     for comat in comat_data:
#         data.append({
#             "stdf_pk": comat.stdf_pk,
#             "awb": comat.awb,
#         })

#     return JsonResponse({
#         "data": data,
#         "draw": draw,
#         "recordsTotal": Comat.objects.count(),  # Total de registros sin filtrar
#         "recordsFiltered": Comat.objects.count(),  # Total de registros después del filtrado (puedes ajustar esto según tus necesidades)
#     })



#OBTIENE LOS RESULTADOS CON MÁS RELACION QUE TIENE LA BUSQUEDA
def obtener_datos_comat(request):
    # Obtén los parámetros enviados por DataTables
    draw = int(request.GET.get('draw', 0))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))  # Número de registros por página
    search_value = request.GET.get('c', '')  # Término de búsqueda

    comat_data = Comat.objects.all()
    
    if search_value:
        comat_data = comat_data.filter(Q(stdf_pk__icontains=search_value) | Q(awb__icontains=search_value)| Q(hawb__icontains=search_value))
        


    # Realizar la consulta teniendo en cuenta la paginación
    comat_data = comat_data[start:start + length]


    # Formatea los datos en un formato compatible con DataTables
    data = []
    for comat in comat_data:
        data.append({
            "stdf_pk": comat.stdf_pk,
            "awb": comat.awb,
            "hawb":comat.hawb,
            "num_manifiesto":comat.num_manifiesto,
            "sum_cif":comat.sum_cif,
            "bodega_fk":comat.bodega_fk.name_bodega,
            "usuario": comat.usuario.username,
            
        })
    
    if search_value:
        records_filtered = Comat.objects.filter(stdf_pk__icontains=search_value).count()
    else:
    # Si no hay término de búsqueda, simplemente cuenta todos los registros
        records_filtered = Comat.objects.count()

    return JsonResponse({
        "data": data,
        "draw": draw,
        "recordsTotal": Comat.objects.count(),  # Total de registros sin filtrar
        "recordsFiltered": records_filtered  # Total de registros después del filtrado (puedes ajustar esto según tus necesidades)
    })

#OBTIENE LOS DATOS PARA REALIZAR EL DETALLE POR LA ID
def detalle_comat(request, stdf_pk):
    detalle_comat = Comat.objects.get(stdf_pk=stdf_pk)    
    return render(request,'detalle_comat.html' , {'detalle_comat':detalle_comat})






#VISTA DE INCOMING QUE VALIDA EL FORMULARIO Y LO GUARDA Y REDIRIGE A LA PAGINA DE INCOMING
@login_required 
def incoming(request):
    get_form_incoming = Incoming.objects.all()
    total_unit_cost = 0
    if request.method == 'POST':
        form_incoming = IncomingForm(request.POST)
        if form_incoming.is_valid():
                if request.user.is_authenticated:
                    # Guarda el formulario de Incoming
                    incoming = form_incoming.save(commit=False)

                    incoming.usuario = request.user

                    total_unit_cost =  incoming.qty * incoming.u_purchase_cost 
                    # Copia el valor de cantidad_extraida a la columna saldo
                    incoming.total_u_purchase_cost = total_unit_cost
                    incoming.saldo = incoming.qty
                    incoming.save()


                    return redirect('/incoming')
                else:
                # Manejo del caso en el que el usuario no está autenticado
                    return HttpResponse("Debes iniciar sesión para realizar esta acción.")
    else:
        form_incoming = IncomingForm()

    context = {
        'form_incoming': form_incoming,
        'get_form_incoming': get_form_incoming, 
    }
    return render(request, 'incoming.html', context)


#GUARDA EL VALOR QUE SE BUSCO Y REDIRIGE A LA PAGINA DE RESULTADOS
def buscar_productos_incoming(request):
    # Obtiene el término de búsqueda del usuario desde la URL
    query_inco = request.GET.get('e', '')
    
    return render(request, 'resultado_busqueda_incoming.html', {'query_inco':query_inco})
    
    # Realiza la búsqueda de productos por id_stdf
    #resultados_part = Incoming.objects.filter(part_number__icontains=query_part)
    # resultados_inc = Incoming.objects.filter(id_stdf=query_inc)


#OBTIENE LOS DATOS RELACIONADOS A LA BUSQUEDA 
def obtener_datos_incoming(request):
    # Obtén los parámetros enviados por DataTables
    draw = int(request.GET.get('draw', 0))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))  # Número de registros por página
    search_value = request.GET.get('e', '')  # Término de búsqueda

    incoming_data = Incoming.objects.all()
    
    if search_value:
        incoming_data = incoming_data.filter(Q(part_number__icontains = search_value) | Q(sn_batch_pk__icontains=search_value))
        


    # Realizar la consulta teniendo en cuenta la paginación
    incoming_data = incoming_data[start:start + length]


    # Formatea los datos en un formato compatible con DataTables
    data = []
    for incoming in incoming_data:
        data.append({
            "sn_batch_pk":incoming.sn_batch_pk,
            "categoria_fk":incoming.categoria_fk.name_categoria,
            "part_number":incoming.part_number,
            "descripcion": incoming.descripcion,
            "qty":incoming.qty,
            "f_vencimiento":incoming.f_vencimiento,
            "usuario": incoming.usuario.username,
            "saldo":incoming.saldo,
        })
    
    if search_value:
        records_filtered = Incoming.objects.filter(part_number__icontains=search_value).count()
    else:
    # Si no hay término de búsqueda, simplemente cuenta todos los registros
        records_filtered = Incoming.objects.count()

    return JsonResponse({
        "data": data,
        "draw": draw,
        "recordsTotal": Incoming.objects.count(),  # Total de registros sin filtrar
        "recordsFiltered": records_filtered  # Total de registros después del filtrado (puedes ajustar esto según tus necesidades)
    })


def detalle_incoming(request, sn_batch_pk):
    detalle_incoming = Incoming.objects.get(sn_batch_pk=sn_batch_pk)    
    return render(request,'detalle_incoming.html' , {'detalle_incoming':detalle_incoming})



#VISTAS DE CONSUMO

#VISTA DE CONSUMO QUE VALIDA EL FORMULARIO Y LO GUARDA Y REDIRIGE A LA VISTA DE CONSUMOS
def consumos(request):
    form_consumos = ConsumosForm()
    if request.method == 'POST':
        form_consumos = ConsumosForm(request.POST)
        if form_consumos.is_valid():
            if request.user.is_authenticated:
                consumo = form_consumos.save(commit=False)  # Guardar el formulario de Consumos sin guardarlo en la base de datos

                consumo.usuario = request.user

                # Obtener el registro correspondiente en Incoming
                incoming = consumo.incoming_fk

                incoming.saldo -=consumo.qty_extraida

                # Actualizar el valor de saldo en el registro de Incoming
                incoming.save()

                # Guardar el registro de Consumos en la base de datos
                consumo.save()

                return redirect('/consumos')
            else:
                # Manejo del caso en el que el usuario no está autenticado
                return HttpResponse("Debes iniciar sesión para realizar esta acción.")  

    context = {
        'form_consumos': form_consumos,
    }
    return render(request, 'consumos.html', context)



   
#GUARDA LA BUSQUEDA QUE SE REALIZO Y REDIRIGE A LA PAGINA DE RESULTADOS DE CONSUMOS
def buscar_productos_consumos(request):
    # Obtiene el término de búsqueda del usuario desde la URL
    query_consu = request.GET.get('t', '')

    return render(request, 'resultado_busqueda_consumos.html', {'query_consu': query_consu})



#OBTIENE LOS DATOS RELACIONADOS A LA BUSQUEDA DE CONSUMOS
def obtener_datos_consumos(request):
    # Obtén los parámetros enviados por DataTables
    draw = int(request.GET.get('draw', 0))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))  # Número de registros por página
    search_value = request.GET.get('t', '')  # Término de búsqueda

    # Inicializa la consulta sin filtrar
    consumos_data = Consumos.objects.all()

    # Realiza la consulta teniendo en cuenta el término de búsqueda en incoming_fk
    if search_value:
            consumos_data = consumos_data.filter(Q(incoming_fk=search_value))

    # Realiza la paginación
    consumos_data = consumos_data[start:start + length]

    # Formatea los datos en un formato compatible con DataTables
    data = []
    for consumo in consumos_data:
        data.append({
            "consumo_pk": consumo.consumo_pk,
            "incoming_fk": consumo.incoming_fk.sn_batch_pk,
            "f_transaccion": consumo.f_transaccion,
            "matricula_aeronave": consumo.matricula_aeronave,
            "orden_consumo": consumo.orden_consumo,
            "qty_extraida": consumo.qty_extraida,
            "usuario": consumo.usuario.username,
        })

    if search_value:
        records_filtered = Consumos.objects.filter(incoming_fk=search_value).count()
    else:
    # Si no hay término de búsqueda, simplemente cuenta todos los registros
        records_filtered = Consumos.objects.count()

    return JsonResponse({
        "data": data,
        "draw": draw,
        "recordsTotal": Consumos.objects.count(),  # Total de registros sin filtrar
        "recordsFiltered": records_filtered,
        
    })

#OBTIENE LOS DATOS PARA REALIZAR EL DETALLE DE CONSUMOS
def detalle_consumos(request, consumo_pk):
    detalle_consumos = Consumos.objects.get(consumo_pk=consumo_pk)   

    incoming_data = detalle_consumos.incoming_fk

    return render(request,'detalle_consumos.html' , {'detalle_consumos':detalle_consumos, 'incoming_data' : incoming_data })






def detalle_inicio(request, stdf_pk):
    draw = int(request.GET.get('draw', 0))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))  # Número de registros por página

    comat_data = Comat.objects.get(stdf_pk=stdf_pk)

    # Tabla de Incoming
    incoming_objects = Incoming.objects.filter(stdf_fk=comat_data)
    incoming_objects_paginated = incoming_objects[start:start + length]

    incoming_data_list = []

    for incoming_obj in incoming_objects_paginated:
        incoming_data_list.append({
            "sn_batch_pk": incoming_obj.sn_batch_pk,
            "part_number": incoming_obj.part_number,
            "f_incoming": incoming_obj.f_incoming,
            "descripcion": incoming_obj.descripcion,
            "po":incoming_obj.po,
            "qty":incoming_obj.qty,
            "u_purchase_cost":incoming_obj.u_purchase_cost,
            "total_u_purchase_cost":incoming_obj.total_u_purchase_cost,
            "f_vencimiento":incoming_obj.f_vencimiento,
            "saldo":incoming_obj.saldo,
            "observaciones":incoming_obj.observaciones,
            "categoria_fk":incoming_obj.categoria_fk.name_categoria,
            "clasificacion_fk":incoming_obj.clasificacion_fk.name_clasificacion,
            "ubicacion_fk":incoming_obj.ubicacion_fk.name_ubicacion,
            "uom_fk":incoming_obj.uom_fk.name_uom,
            "owner_fk":incoming_obj.owner_fk.name_owner,
            "condicion_fk":incoming_obj.condicion_fk.name_condicion,
            "ficha_fk":incoming_obj.ficha_fk.name_ficha,
    })

    incoming_records_total = incoming_objects.count()
    incoming_records_filtered = incoming_records_total  # Opcionalmente, puedes aplicar filtros aquí

    # Tabla de Consumos
    consumos_objects = Consumos.objects.filter(incoming_fk__in=incoming_objects)
    consumos_objects_paginated = consumos_objects[start:start + length]

    consumos_data_list = []

    for consumos_obj in consumos_objects_paginated:
        consumos_data_list.append({
                "incoming_fk":consumos_obj.incoming_fk.sn_batch_pk,
                "orden_consumo":consumos_obj.orden_consumo,
                "f_transaccion":consumos_obj.f_transaccion,
                "qty_extraida":consumos_obj.qty_extraida,
                "matricula_aeronave":consumos_obj.matricula_aeronave,
                "observaciones":consumos_obj.observaciones,
        })

    consumos_records_total = consumos_objects.count()
    consumos_records_filtered = consumos_records_total  # Opcionalmente, puedes aplicar filtros aquí

    data = {
        "draw": draw,
        "recordsTotal": incoming_records_total,
        "recordsFiltered": incoming_records_filtered,
        "comat_data": {
                "stdf_pk": comat_data.stdf_pk,
                "awb": comat_data.awb,
                "hawb": comat_data.hawb,
                "num_manifiesto": comat_data.num_manifiesto,
                "sum_cif": comat_data.sum_cif,
                "corr_interno" : comat_data.corr_interno,
                "psc": comat_data.pcs,
                "peso": comat_data.peso,
                "f_control": comat_data.f_control,
                "f_manifiesto": comat_data.f_manifiesto,
                "f_recepcion": comat_data.f_recepcion,
                "f_stdf": comat_data.f_stdf,
                "fob": comat_data.fob,
                "flete": comat_data.flete,
                "seguro": comat_data.seguro,
                "bodega_fk": comat_data.bodega_fk.name_bodega,
                "origen_fk": comat_data.origen_fk.name_origen,
                "estado_fk": comat_data.estado_fk.estado,
            },
        "incoming_data": incoming_data_list,
        "consumos_data": consumos_data_list,
    }

    if request.META.get("HTTP_X_REQUESTED_WITH") == "XMLHttpRequest":
        return JsonResponse(data)
    else:
        # Si no es una solicitud AJAX, renderiza una plantilla HTML
        return render(request, 'detalle_inicio.html', {
            "comat_data": comat_data,
            "incoming_data_list": incoming_data_list,
            "incoming_recordsTotal": incoming_records_total,
            "incoming_recordsFiltered": incoming_records_filtered,
            "consumos_data": consumos_data_list,
            "consumos_recordsTotal": consumos_records_total,
            "consumos_recordsFiltered": consumos_records_filtered,
        })



def exportar_excel_incoming(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="datos.xlsx"'

    wk = Workbook()
    ws = wk.active



    ancho_columna = 3.55

    for columna in range(1, 26):
        # Establece el ancho de la columna
        ws.column_dimensions[ws.cell(row=1, column=columna).column_letter].width = ancho_columna


    filas_altura_6 = [3 ,6 ,12 , 16 , 18 , 20 , 22 , 24, 28 , 68]  # Cambia estos valores a las filas que desees

    # Define la altura predeterminada para las demás filas (12 en este caso)
    altura_predeterminada = 12

    # Recorre todas las filas de la hoja
    for fila in ws.iter_rows(min_row=1, max_row=73):
        numero_fila = fila[0].row  # Obtiene el número de fila de la primera celda de la fila
        if numero_fila in filas_altura_6:
            ws.row_dimensions[numero_fila].height = 6
        else:
            ws.row_dimensions[numero_fila].height = altura_predeterminada



    formato_fuente = Font(name='tahoma', size=8) 

    for fila in ws.iter_rows(min_row=1, max_row=74, min_col=1, max_col=25):  # 25 es la columna Y
        for celda in fila:
            celda.font = formato_fuente


    border = Border(
        left=Side(border_style="thin"),   # Bordes izquierdos
        right=Side(border_style="thin"),  # Bordes derechos
        top=Side(border_style="thin"),    # Bordes superiores
        bottom=Side(border_style="thin")  # Bordes inferiores
    )


    coordenadas_excepciones = ['R1:T2']

    alineacion = Alignment(horizontal='center', vertical='center')

    rango_celdas = 'A1:Y72'

    ws.merge_cells('R1:T2')
    ws['R1']  = 'RCV N°'

    ws.merge_cells('U1:Y2')

    ws.merge_cells('H5:Q5')
    ws['H5']  = 'INSPECCION DE RECEPCION'

    ws.merge_cells('A8:G8')
    ws['A8']  = 'Descripcion'
    ws.merge_cells('A9:G11')#Campo Relleno de BBDD 


    ws.merge_cells('H8:O8')
    ws['H8']  = 'N° de Parte/PN'
    ws.merge_cells('H9:O11')#Campo Relleno de BBDD 

    ws.merge_cells('P8:S8')
    ws['P8']  = 'Modelo'
    ws.merge_cells('P9:S11')#Campo Relleno de BBDD 

    ws.merge_cells('T8:Y8')
    ws['T8']  = 'N° de Serie/SN'
    ws.merge_cells('T9:Y11')#Campo Relleno de BBDD 

    #FALTA CAMBIAR ESPACIO FILA 12

    ws.merge_cells('A13:C13')
    ws['A13']  = 'Overhaul'

    ws.merge_cells('A15:C15')
    ws['A15']  = 'Reparado'

    ws.merge_cells('E13:G13')
    ws['E13']  = 'Chequeado'

    ws.merge_cells('E15:G15')
    ws['E15']  = 'Testeado'

    ws.merge_cells('I13:K13')
    ws['I13']  = 'Otro'

    ws.merge_cells('I15:K15')
    ws['I15']  = 'Nuevo'

    ws.merge_cells('M13:P13')
    ws['M13']  = 'Calibración'

    ws.merge_cells('Q13:T14')#Campo Relleno de BBDD 
    ws.merge_cells('Q15:T15')
    ws['Q15']  = 'Fecha Recepción'


    ws.merge_cells('U13:W15')
    ws['U13']  = 'Cantidad'
    ws.merge_cells('X13:Y15')#Campo Relleno de BBDD 

    ws.merge_cells('A17:G17')
    ws['A17']  = 'Proveedor del Producto'
    ws.merge_cells('H17:P17') #Campo Relleno de BBDD 


    ws.merge_cells('Q17:S17')
    ws['Q17']  = 'OC / PO N°'
    ws.merge_cells('T17:Y17') #Campo Relleno de BBDD 

    ws.merge_cells('A19:G19')
    ws['A19']  = 'Taller/Cia. Reparadora'
    ws.merge_cells('H19:P19') #Campo Relleno de BBDD 

    ws.merge_cells('Q19:S19')
    ws['Q19']  = 'RO N°'
    ws.merge_cells('T19:Y19') #Campo Relleno de BBDD 

    ws.merge_cells('A21:G21')
    ws['A21']  = 'Trabajo Solicitado'
    ws.merge_cells('H21:P21') #Campo Relleno de BBDD 

    ws.merge_cells('Q21:S21')
    ws['Q21']  = 'WO N°'
    ws.merge_cells('T21:Y21') #Campo Relleno de BBDD 


    ws.merge_cells('A23:G23')
    ws['A23']  = 'Propiedad de'
    ws.merge_cells('H23:Y23') #Campo Relleno de BBDD 

    ws.merge_cells('A25:G25')
    ws['A25']  = 'Check Periodica (Si Aplica)'
    ws.merge_cells('H25:N25') #Campo Relleno de BBDD 



    ws.merge_cells('O25:R25')
    ws['O25']  = 'Shel life Due'
    ws.merge_cells('S25:Y25') #Campo Relleno de BBDD 

    #ENCABEZADO DE LA TABLA
    ws['A27'] = 'Item'
    ws.merge_cells('B27:S27')
    ws['B27']  = 'Materia'
    ws.merge_cells('T27:V27') 
    ws['T27']  = 'SI'
    ws.merge_cells('W27:Y27') 
    ws['W27']  = 'NO'

    #CUERPO DE LA TABLA
    ws['A29'] = '1'
    ws.merge_cells('B29:S29')
    ws['B29'] = 'Producto conforme a lo indicado en la lista de Embarque (Packing List)'
    ws.merge_cells('T29:V29')
    ws.merge_cells('W29:Y29')

    ws['A30'] = '2'
    ws.merge_cells('B30:S30')
    ws['B30'] = 'Factura del Proveedor conforme a la orden de compra o solicitud de trabajo (Invoice)'
    ws.merge_cells('T30:V30')
    ws.merge_cells('W30:Y30')

    ws['A31'] = '3'
    ws.merge_cells('B31:S31')
    ws['B31'] = 'Cartilla/Orden de Trabajo, Cartilla de prueba, Si corresponde, del taller que repara'
    ws.merge_cells('T31:V31')
    ws.merge_cells('W31:Y31')

    ws['A32'] = '4'
    ws.merge_cells('B32:S32')
    ws['B32'] = 'Producto sin daños visibles (Inspeccion Visual)'
    ws.merge_cells('T32:V32')
    ws.merge_cells('W32:Y32')


    ws['A33'] = '5'
    ws.merge_cells('B33:S33')
    ws['B33'] = 'Producto protegido en embalaje apropiado'
    ws.merge_cells('T33:V33')
    ws.merge_cells('W33:Y33')

    ws['A34'] = '6'
    ws.merge_cells('B34:S34')
    ws['B34'] = 'Placa de identificacion del componente'
    ws.merge_cells('T34:V34')
    ws.merge_cells('W34:Y34')

    ws['A35'] = '7'
    ws.merge_cells('B35:S35')
    ws['B35'] = 'Documentacion Técnica completa requerida por reglamentacion (Trazabilidad)'
    ws.merge_cells('T35:V35')
    ws.merge_cells('W35:Y35')

    ws['A36'] = '8'
    ws.merge_cells('B36:S36')
    ws['B36'] = 'Formulario FAA 8130-3'
    ws.merge_cells('T36:V36')
    ws.merge_cells('W36:Y36')

    ws['A37'] = '9'
    ws.merge_cells('B37:S37')
    ws['B37'] = 'Formulario EASA Form One o JAA Form One o CAA Form One'
    ws.merge_cells('T37:V37')
    ws.merge_cells('W37:Y37')

    ws['A38'] = '10'
    ws.merge_cells('B38:S38')
    ws['B38'] = 'Formulario DGAC Chile 8130-3'
    ws.merge_cells('T38:V38')
    ws.merge_cells('W38:Y38')

    ws['A39'] = '11'
    ws.merge_cells('B39:S39')
    ws['B39'] = 'Formulario ANAC Argentina 8130-3																	'
    ws.merge_cells('T39:V39')
    ws.merge_cells('W39:Y39')

    ws['A40'] = '12'
    ws.merge_cells('B40:S40')
    ws['B40'] = 'Placa o Etiqueta para Herramientas o equipos con calibracion'
    ws.merge_cells('T40:V40')
    ws.merge_cells('W40:Y40')

    ws['A41'] = '13'
    ws.merge_cells('B41:M41')
    ws['B41'] = 'Certificado de Calibracion en laboratorio reconocido por el estado local'
    ws['N41'] = 'N°'
    ws.merge_cells('O41:S41')
    ws.merge_cells('T41:V41')
    ws.merge_cells('W41:Y41')


    ws['A42'] = '14'
    ws.merge_cells('B42:S42')
    ws['B42'] = 'Materiales con vida limite (Verificacion de Shelf life data y MSDS)'
    ws.merge_cells('T42:V42')
    ws.merge_cells('W42:Y42')

    ws['A43'] = '15'
    ws.merge_cells('B43:J43')
    ws['B43'] = 'Certificado de flamabilidad, si corresponde'
    ws.merge_cells('L43:S43')#CAMPO LLENADO POR BBDD
    ws['K43'] = 'N°'
    ws.merge_cells('T43:V43')
    ws.merge_cells('W43:Y43')

    ws['A44'] = '16'
    ws.merge_cells('B44:J44')
    ws['B44'] = 'Certificado de conformidad  y/o Analisis'
    ws.merge_cells('L44:S44')#CAMPO LLENADO POR BBDD
    ws['K44'] = 'N°'
    ws.merge_cells('T44:V44')
    ws.merge_cells('W44:Y44')

    ws['A45'] = '17'
    ws.merge_cells('B45:J45')
    ws['B45'] = 'Numero de lote de fabricacion, si corresponde'
    ws.merge_cells('L45:S45')#CAMPO LLENADO POR BBDD
    ws['K45'] = 'N°'
    ws.merge_cells('T45:V45')
    ws.merge_cells('W45:Y45')

    ws['A46'] = '18'
    ws.merge_cells('B46:E46')
    ws['B46'] = 'TSO / TSN (Si Aplica)'

    ws.merge_cells('G45:H45')#CAMPO LLENADO POR BBDD
    ws['F46'] = 'TSN'

    ws.merge_cells('I46:K46')
    ws['I46'] = 'TSO'

    ws.merge_cells('M46:O46')
    ws['L46'] = 'CSN'

    ws.merge_cells('Q46:S46')
    ws['P46'] = 'CSO'

    ws.merge_cells('T46:V46')
    ws.merge_cells('W46:Y46')

    ws['A47'] = '19'
    ws.merge_cells('B47:S47')
    ws['B47'] = 'Material Safety Data Sheet'
    ws.merge_cells('T47:V47')
    ws.merge_cells('W47:Y47')

    ws['A48'] = '20'
    ws.merge_cells('B48:S48')
    ws['B48'] = 'Material con restriccion bajo el programa ESD'
    ws.merge_cells('T48:V48')
    ws.merge_cells('W48:Y48')

    ws['A49'] = '21'
    ws.merge_cells('B49:S49')
    ws['B49'] = 'Material con restriccion de almacenamiento (Hielo Seco)'
    ws.merge_cells('T49:V49')
    ws.merge_cells('W49:Y49')

    ws['A50'] = '22'
    ws.merge_cells('B50:I50')
    ws['B50'] = 'Cartilla Mantencion CMA Autorizado'

    ws.merge_cells('K50:S50')
    ws['J50'] = 'N°'

    ws.merge_cells('T50:V50')
    ws.merge_cells('W50:Y50')

    ws.merge_cells('A52:Y52')
    ws['A52'] = 'Observaciones'
    ws.merge_cells('A53:Y56')

    ws.merge_cells('A58:Y58')
    ws['A58'] = 'Observaciones del Producto Recepcionado'

    ws.merge_cells('A59:D59')
    ws['A59'] = 'OC, PO OR RO N°'
    ws.merge_cells('E59:Y59')


    ws.merge_cells('A60:D60')
    ws['A60'] = 'BATCH'
    ws.merge_cells('E60:Y60')

    ws.merge_cells('A61:D61')
    ws['A61'] = 'STDF'
    ws.merge_cells('E61:Y61')

    ws.merge_cells('A62:D62')
    ws['A62'] = 'AWB'
    ws.merge_cells('E62:Y62')

    ws.merge_cells('A63:D64')
    ws['A63'] = 'UBICACION'
    ws.merge_cells('E63:Y64')

    ws.merge_cells('A65:D65')
    ws['A65'] = 'ACEPTADO'

    ws.merge_cells('E65:F65')
    ws['E65'] = 'SI'
    ws.merge_cells('G65:H65')

    ws.merge_cells('I65:J65')
    ws['I65'] = 'NO'
    ws.merge_cells('K65:L65')


    ws.merge_cells('A67:H67')
    ws['I65'] = 'Nombre'

    ws.merge_cells('I67:P67')
    ws['I67'] = 'N° Licencia'

    ws.merge_cells('Q67:Y67')
    ws['Q67'] = 'Firma'

    ws.merge_cells('A69:H71')
    ws['A69'] = '' #NOMBRE DE USUARIO
    
    ws.merge_cells('I69:P71')
    ws['I67'] = 'INCOMING INSPECTION'

    ws.merge_cells('Q69:Y71') #FIRMA DEL USUARIO 













    #ws.merge_cells('')
    #ws['']  = ''
    #ws.merge_cells(':') #Campo Relleno de BBDD 


    coordenadas_combinadas = ws.merged_cells.ranges
    for rango in coordenadas_combinadas:
        for row in ws.iter_rows(min_row=rango.min_row, max_row=rango.max_row,
                                    min_col=rango.min_col, max_col=rango.max_col):
            for cell in row:
                if cell.coordinate not in coordenadas_excepciones:
                    cell.border = border






    for fila in ws.iter_rows(min_row=1, max_row=72, min_col=1, max_col=25):
        for cell in fila:
            cell.alignment = alineacion



    # Configura la escala de la hoja para ajustarla a una página
    ws.page_setup.fitToPage = True

    # Configura los márgenes de la página para reducirlos si es necesario
    ws.page_margins.left = 0.25  # Márgen izquierdo
    ws.page_margins.right = 0.30  # Márgen derecho
    ws.page_margins.top = 0.25  # Márgen superior
    ws.page_margins.bottom = 0.25  # Márgen inferior

    

    wk.save(response)
    return response

















































# ## Test
# def obtener_datos_stdf_incoming(request):

#     term = request.GET.get('q', '')

#     stdf_data = Comat.objects.filter(stdf_pk__icontains=term).values('stdf_pk')[:20]

#     # stdf_data = Comat.objects.all().values('stdf_pk')
#     stdf_list = list(stdf_data)
    
#     # Convierte la lista de diccionarios a una lista de objetos JSON
#     stdf_json = [{'stdf_pk': item['stdf_pk']} for item in stdf_list]
    
#     return JsonResponse({'stdf_data': stdf_json}, safe=False)





#VISTA DASHBOARD
def dashboard(request):
    comat_data = Comat.objects.all().order_by()[:10]
    incoming_data = Incoming.objects.all()
    context = {
        'comat_data':comat_data,
        'incoming_data':incoming_data

    }

    return render(request, 'dashboard.html', context)